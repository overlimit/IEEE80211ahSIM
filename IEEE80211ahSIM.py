
# coding: utf-8

# In[ ]:

from __future__ import division
class Node_common:
    'Common base class for all nodes'

    DEVICE = "node"
    
    from random import randint
    
    def __init__(self, x, y, packetLength, samplingRate, ID):
        self.x = x
        self.y = y
        self.packetLength = packetLength
        self.samplingRate = samplingRate
        self.nodeInRange = []
        self.channelBusyCount = 0
        self.state = Global.HALT
        self.timeToNextTask = Global.maxTime +1
        self.nav = Global.nav
        self.checkACK = 0
        self.backoffStage = Global.contentionWindow
        self.backoffTime = 0
        self.tranTime = fix(((packetLength * 8) / Global.slotTime) / Global.dataRate)
        self.queuingData = 0
        self.collision = 0
        self.ID = ID
        self.transmittTimes = 0
        self.expectedUsingTimePerBeacon = round(((self.tranTime * (samplingRate / 60)) * Global.beacon_sec), 5)
        self.transmmitSuccessPerCycle = 0
        self.simultaneouslyCount = 0
        self.totalSamplingTime = 0
        self.tryAgainFlag = 0
        Global.nodeCounts += 1
        
    def sendPacket(self):
        '''
        to AP: send request
        to other nodes: message "transmitting"
        '''
        Global.packetCount += 1
        Global.channelUser += 1
        self.transmittTimes += 1
        self.timeToNextTask = Global.currentTime + self.tranTime
        self.state = Global.sendPacket
        #print "Node: %d, send packet, raw time: %d, period time: %d" % (self.ID, Global.rawInterval, Global.holdingPeriodInterval)
        for nodeSendIndex in range(len(self.nodeInRange)):
            self.nodeInRange[nodeSendIndex].receivePacket()
        return self
            
    
    def receivePacket(self):
        'Carrier sense'
        self.channelBusyCount += 1
        if (self.state == Global.waitDIFS) or (self.state == Global.backoff):
            if self.state == Global.backoff:
                self.backoffTime = self.backoffTime - (Global.currentTime - self.backoffStartTime)
            self.state = Global.carrierSense
            #print "Node: %d, carrier sense" % self.ID
            self.timeToNextTask = Global.maxTime +1
        if self.channelBusyCount >1:
            self.nav = 0
            self.checkACK = 0
            
        
    def changeState(self):
        "change state when it's the node's turn"
        if self.state == Global.sendPacket:
            return self.toWaitResponse()
        elif self.state == Global.waitResponse:
            return self.toCheckACK()
        elif self.state == Global.waitDIFS:
            return self.toBackoff()
        elif self.state == Global.backoff:
            return self.readyToWork()
    
    def toWaitResponse(self):
        self.state = Global.waitResponse
        #print "Node: %d, wait response" % self.ID
        self.timeToNextTask = Global.currentTime + Global.nav
        for userDecreaseIndex in range(len(self.nodeInRange)):
            self.nodeInRange[userDecreaseIndex].channelUserDecrease()
        Global.channelUser -= 1 
            
    def channelUserDecrease(self):
        self.channelBusyCount -= 1
        if (self.channelBusyCount < 1) and (self.state == Global.carrierSense):
            #print "Node: %d, carrier sense over,  backoff time: %d" % (self.ID, self.backoffTime)
            self.state = Global.waitDIFS
            self.timeToNextTask = Global.currentTime + self.nav + Global.DIFS
            self.nav = Global.nav
        
        
    def toCheckACK(self):
        'check whether ACK is transmitted successfully'
        #print "Node: %d, check ACK" % self.ID
        'if failed => toBackoff'
        if self.checkACK == 1:
            self.backoffStage = Global.contentionWindow
            #print "Node: %d, transmitt success!" % self.ID
            'transmitted successfully'
            Global.success += 1
            self.transmmitSuccessPerCycle += 1
            self.queuingData -= 1
            if self.queuingData > 0:
                if self.channelBusyCount > 0:
                    self.state = Global.carrierSense
                    self.timeToNextTask = Global.maxTime +1
                else:
                    self.state = Global.waitDIFS
                    self.timeToNextTask = Global.currentTime + Global.DIFS
            else:
                self.state = Global.HALT
                self.timeToNextTask = Global.maxTime +1
            self.checkACK = 0
        elif self.checkACK == 0:
            #print "Node: %d, collision occured" % self.ID
            self.collision += 1
            Global.collisionTimes += 1
            if self.channelBusyCount > 0:
                self.state = Global.carrierSense
                self.timeToNextTask = Global.maxTime +1
            else:
                self.state = Global.waitDIFS
                self.timeToNextTask = Global.currentTime + Global.DIFS
        
    def readyToWork(self):
        self.backoffTime = 0
        if Global.currentTime < Global.holdingPeriodInterval:
            self.state = Global.READY
        else:
            self.state = Global.HALT
            self.timeToNextTask = Global.maxTime +1
        
        
    def toBackoff(self):
        #self.nav = 0
        self.state = Global.backoff
        if self.backoffTime == 0:
            self.backoffStage += 1
            self.backoffTime = Global.backoffSeed.randint(1, 2 ** min(self.backoffStage, 10))
            self.timeToNextTask = Global.currentTime + self.backoffTime
        else:
            self.timeToNextTask = Global.currentTime + self.backoffTime
        self.backoffStartTime = Global.currentTime
        #print "Node: %d, wait backoff: %d" % (self.ID, self.backoffTime)
        
    def dataInterval(self):
        self.queuingData += 1
        self.totalSamplingTime += 1
        if self.transmmitSuccessPerCycle > 0:
            self.simultaneouslyCount += 1 
        self.dataIntervalTime = Global.currentTime
        self.transmmitSuccessPerCycle = 0
        self.tryAgainFlag = 0  
        #new data is arrived
        '''if self.state == Global.HALT:
            self.state = Global.waitDIFS
            self.timeToNextTask = Global.currentTime + Global.DIFS'''
        
    def wakeUp(self):
        if self.queuingData > 0:
            self.state = Global.waitDIFS
            self.timeToNextTask = Global.currentTime + Global.DIFS
            Global.competitorCount += 1
            Global.extryCompetitorCount += self.tryAgainFlag
        #print "node: %d, state: %d" % (self.ID, self.state)
            
    def goSleep(self):
        self.state = Global.HALT
        self.timeToNextTask = Global.maxTime +1
        if self.queuingData > 0:
            self.tryAgainFlag = 1
        else:
            self.tryAgainFlag = 0
        
        
    def calcRange(self, others):
        if (((self.x- others.x) ** 2) + ((self.y - others.y) ** 2)) <= 1000000:
            self.nodeInRange.append(others)
            others.nodeInRange.append(self)
    
    def displayNodesInRange(self):
        for x in range(len(self.nodeInRange)):
            print self.nodeInRange[x].displayStatus()
        
    def displayStatus(self):
        print "x: %d, y: %d,\nType: %s, group: %d" %(self.x, self.y, self.device, self.group)
                    
    def displayNodeCounts(self):
        print "Total nodes: %d" % Global.nodeCounts
        


# In[ ]:

from __future__ import division
class Node_scheduled:
    'Common base class for all nodes'

    DEVICE = "node"
    
    from random import randint
    
    def __init__(self, x, y, packetLength, samplingRate, ID):
        self.x = x
        self.y = y
        self.packetLength = packetLength
        self.samplingRate = samplingRate
        self.nodeInRange = []
        self.channelBusyCount = 0
        self.state = Global.HALT
        self.timeToNextTask = Global.maxTime +1
        self.nav = Global.nav
        self.checkACK = 0
        self.backoffStage = Global.contentionWindow
        self.backoffTime = 0
        self.tranTime = fix(((packetLength * 8) / Global.slotTime) / Global.dataRate)
        self.queuingData = 0
        self.collision = 0
        self.ID = ID
        self.transmittTimes = 0
        self.expectedUsingTimePerBeacon = round(((self.tranTime * (samplingRate / 60)) * Global.beacon_sec), 5)
        self.beaconCount = 0
        self.transmmitSuccessPerCycle = 0
        self.simultaneouslyCount = 0
        self.totalSamplingTime = 0
        self.tryAgainFlag = 0
        Global.nodeCounts += 1
        
    def sendPacket(self):
        '''
        to AP: send request
        to other nodes: message "transmitting"
        '''
        Global.packetCount += 1
        Global.channelUser += 1
        self.transmittTimes += 1
        self.timeToNextTask = Global.currentTime + self.tranTime
        self.state = Global.sendPacket
        #print "Node: %d, send packet, raw time: %d, period time: %d" % (self.ID, Global.rawInterval, Global.holdingPeriodInterval)
        for nodeSendIndex in range(len(self.nodeInRange)):
            self.nodeInRange[nodeSendIndex].receivePacket()
        return self
            
    
    def receivePacket(self):
        'Carrier sense'
        self.channelBusyCount += 1
        if (self.state == Global.waitDIFS) or (self.state == Global.backoff):
            if self.state == Global.backoff:
                self.backoffTime = self.backoffTime - (Global.currentTime - self.backoffStartTime)
            self.state = Global.carrierSense
            #print "Node: %d, carrier sense" % self.ID
            self.timeToNextTask = Global.maxTime +1
        if self.channelBusyCount >1:
            self.nav = 0
            self.checkACK = 0
            
        
    def changeState(self):
        "change state when it's the node's turn"
        if self.state == Global.sendPacket:
            return self.toWaitResponse()
        elif self.state == Global.waitResponse:
            return self.toCheckACK()
        elif self.state == Global.waitDIFS:
            return self.toBackoff()
        elif self.state == Global.backoff:
            return self.readyToWork()
    
    def toWaitResponse(self):
        self.state = Global.waitResponse
        #print "Node: %d, wait response" % self.ID
        self.timeToNextTask = Global.currentTime + Global.nav
        for userDecreaseIndex in range(len(self.nodeInRange)):
            self.nodeInRange[userDecreaseIndex].channelUserDecrease()
        Global.channelUser -= 1 
            
    def channelUserDecrease(self):
        self.channelBusyCount -= 1
        if (self.channelBusyCount < 1) and (self.state == Global.carrierSense):
            #print "Node: %d, carrier sense over,  backoff time: %d" % (self.ID, self.backoffTime)
            self.state = Global.waitDIFS
            self.timeToNextTask = Global.currentTime + self.nav + Global.DIFS
            self.nav = Global.nav
        
        
    def toCheckACK(self):
        'check whether ACK is transmitted successfully'
        #print "Node: %d, check ACK" % self.ID
        'if failed => toBackoff'
        if self.checkACK == 1:
            self.backoffStage = Global.contentionWindow
            #print "Node: %d, transmitt success!" % self.ID
            'transmitted successfully'
            Global.success += 1
            self.transmmitSuccessPerCycle += 1
            self.queuingData -= 1
            if self.queuingData > 0:
                if self.channelBusyCount > 0:
                    self.state = Global.carrierSense
                    self.timeToNextTask = Global.maxTime +1
                else:
                    self.state = Global.waitDIFS
                    self.timeToNextTask = Global.currentTime + Global.DIFS
            else:
                self.state = Global.HALT
                self.timeToNextTask = Global.maxTime +1
            self.checkACK = 0
        elif self.checkACK == 0:
            #print "Node: %d, collision occured" % self.ID
            self.collision += 1
            Global.collisionTimes += 1
            if self.channelBusyCount > 0:
                self.state = Global.carrierSense
                self.timeToNextTask = Global.maxTime +1
            else:
                self.state = Global.waitDIFS
                self.timeToNextTask = Global.currentTime + Global.DIFS
        
    def readyToWork(self):
        self.backoffTime = 0
        if Global.currentTime < Global.holdingPeriodInterval:
            self.state = Global.READY
        else:
            self.state = Global.HALT
            self.timeToNextTask = Global.maxTime +1
        
        
    def toBackoff(self):
        #self.nav = 0
        self.state = Global.backoff
        if self.backoffTime == 0:
            self.backoffStage += 1
            self.backoffTime = Global.backoffSeed.randint(1, 2 ** min(self.backoffStage, 10))
            self.timeToNextTask = Global.currentTime + self.backoffTime
        else:
            self.timeToNextTask = Global.currentTime + self.backoffTime
        self.backoffStartTime = Global.currentTime
        #print "Node: %d, wait backoff: %d" % (self.ID, self.backoffTime)
        
    def dataInterval(self):
        self.queuingData += 1
        self.totalSamplingTime += 1
        if self.transmmitSuccessPerCycle > 0:
            self.simultaneouslyCount += 1 
        self.dataIntervalTime = Global.currentTime
        self.transmmitSuccessPerCycle = 0
        self.tryAgainFlag = 0  
        self.beaconCount = 0   
        #new data is arrived
        '''if self.state == Global.HALT:
            self.state = Global.waitDIFS
            self.timeToNextTask = Global.currentTime + Global.DIFS'''
        
    def wakeUp(self):
        if self.queuingData > 0:
            self.beaconCount -= 1
            if self.beaconCount < 1:
                self.state = Global.waitDIFS
                self.timeToNextTask = Global.currentTime + Global.DIFS
                Global.competitorCount += 1
                Global.extryCompetitorCount += self.tryAgainFlag
        #print "node: %d, state: %d" % (self.ID, self.state)
            
    def goSleep(self):
        self.state = Global.HALT
        self.timeToNextTask = Global.maxTime +1
        if self.queuingData > 0:
            self.tryAgainFlag = 1
        else:
            self.tryAgainFlag = 0
        if self.beaconCount < 1:
            if self.queuingData > 0:
                self.beaconCount = Global.beaconCountSeed.randint(0, fix(((60 / self.samplingRate) * Global.beaconPerSec)))
        
    def calcRange(self, others):
        if (((self.x- others.x) ** 2) + ((self.y - others.y) ** 2)) <= 1000000:
            self.nodeInRange.append(others)
            others.nodeInRange.append(self)
    
    def displayNodesInRange(self):
        for x in range(len(self.nodeInRange)):
            print self.nodeInRange[x].displayStatus()
        
    def displayStatus(self):
        print "x: %d, y: %d,\nType: %s, group: %d" %(self.x, self.y, self.device, self.group)
                    
    def displayNodeCounts(self):
        print "Total nodes: %d" % Global.nodeCounts
        


# In[ ]:

#from __future__ import division
class AP:
    'common access point'
    #DIVICE = "access point"
    
    def __init__(self, x, y, numOfGroup):
        self.x = x
        self.y = y
        self.group = []
        self.groupWeight = []
        self.expectedGroupWeight = []
        self.groupUtilization = []
        self.incrementalGain = []
        self.expectedGroupUtilization = []
        self.channelBusyCount = 0
        self.state = Global.IDLE
        self.timeToNextTask = Global.maxTime +1
        self.nodeInRange = []
        
        self.DEVICE = "access point"
        for groups in range(numOfGroup):
            self.group.append([self])
            self.groupWeight.append(1)
            self.expectedGroupWeight.append(1)
            self.groupUtilization.append(0)
            self.expectedGroupUtilization.append(0)
            self.incrementalGain.append(0)
            
            
            '''
            送封包當下不收封包
            '''
    def receivePacket(self, node):
        'receivePacket'
        self.channelBusyCount += 1
        if self.state != Global.sendPacket:
            if self.channelBusyCount == 1:
                self.state = Global.carrierSense
                self.timeToNextTask = Global.currentTime + node.tranTime + Global.SIFS
                self.respondTarget = node
                #print "AP response Target: %d" % node.ID
            else:
                self.respondTarget = 0
                #print "prefer collision"
                self.timeToNextTask = max(self.timeToNextTask, (Global.currentTime + node.tranTime + Global.SIFS))
        #print "AP channel busy count: %d" % self.channelBusyCount
            
    def sendPacket(self):
        'send ACK'
        #print "AP send ACK to %d" % self.respondTarget.ID
        self.timeToNextTask = Global.currentTime + Global.ACKTime
        self.state = Global.sendPacket
        for APSendRangeIndex in range(len(self.nodeInRange)):
            self.nodeInRange[APSendRangeIndex].receivePacket()
        self.respondTarget.checkACK = 1
        self.respondTarget = 0
        Global.channelUser += 1
        
    def changeState(self):
        self.channelBusyCount = 0
        if self.state == Global.sendPacket:
            return self.toIDLE()
        elif self.state == Global.carrierSense:
            return self.checkRespondTarget()
    
    def processEnd(self):
        print "process end"
        return 0
    
    def toIDLE(self):
        #self.channelBusyCount = 0
        if self.state == Global.sendPacket:
            Global.channelUser -= 1
        self.respondTarget = 0
        self.state = Global.IDLE
        self.timeToNextTask = Global.maxTime +1
        for APRangeIndex in range(len(self.nodeInRange)):
            self.nodeInRange[APRangeIndex].channelBusyCount -= 1
            if (self.nodeInRange[APRangeIndex].channelBusyCount < 1) and (self.nodeInRange[APRangeIndex].state == Global.carrierSense):
                #print "Node: %d, carrier sense over,  backoff time: %d" % (self.nodeInRange[x].ID, self.nodeInRange[x].backoffTime)
                self.nodeInRange[APRangeIndex].state = Global.waitDIFS
                self.nodeInRange[APRangeIndex].timeToNextTask = Global.currentTime + Global.DIFS
                self.nodeInRange[APRangeIndex].nav = Global.nav
        
    def checkRespondTarget(self):
        #self.channelBusyCount = 0
        if self.respondTarget == 0:
            self.toIDLE()
        else:
            self.state = Global.READY       
        
    def calcRange(self, node):
        self.nodeInRange.append(node)

    def displayStatus(self):
        print "x: %d, y: %d,\nType: %s, number of group: %d" %(self.x, self.y, self.device, len(self.group))
        
    def randomGrouping(self, points):
        for randomGrpIndex in range(1, len(points)):
            choiceGroup = self.group[Global.groupingSeed.randint(0, len(self.group))]
            while(len(choiceGroup) >= fix(8192 / len(self.group))):
                choiceGroup = Global.groupingSeed.choice(self.group)
            choiceGroup.append(points[randomGrpIndex])
    
    '''
    def loadBalancedGrouping(self, points):
        for x in range(1, len(points)):
            self.recommandGroup = []
            self.updateGain(points[x])
            #print self.incrementalGain
            for y in range(len(self.group)):
                if self.incrementalGain[y] == max(self.incrementalGain):
                    if len(self.group[y]) < fix(8192 / len(self.group)):
                        self.recommandGroup.append(self.group[y])
            
            self.appendGroup = choice(self.recommandGroup)
            self.appendGroup.append(points[x])
            for y in range(len(self.group)):
                if self.group[y] == self.appendGroup:
                    self.groupWeight[y] = self.expectedGroupWeight[y]
                    self.groupUtilization[y] = self.expectedGroupUtilization[y]
            
    def updateGain(self, node):
        for x in range(len(self.group)):
            self.incrementalGain[x] = 0
            if len(self.group[x]) < 2:
                self.expectedGroupWeight[x] = 1
            else:
                self.expectedGroupWeight[x] = 0
            self.expectedGroupUtilization[x] = 0
            for y in range(1, len(self.group[x])):
                self.expectedGroupWeight[x] = (self.expectedGroupWeight[x] + (self.group[x][y].samplingRate * Global.beacon_sec))
                self.expectedGroupUtilization[x] = (self.expectedGroupUtilization[x] + self.group[x][y].expectedUsingTimePerBeacon)
            self.expectedGroupWeight[x] = (self.expectedGroupWeight[x] + (node.samplingRate * Global.beacon_sec))
            self.expectedGroupWeight[x] = round((1 / self.expectedGroupWeight[x]), 5)
            self.expectedGroupUtilization[x] = round(((self.expectedGroupUtilization[x] + node.expectedUsingTimePerBeacon) / (Global.beacon_sec / Global.group)), 5)
            #self.expectedGroupUtilization[x] = min(self.expectedGroupUtilization, Global.rawTime)
            self.incrementalGain[x] = (self.expectedGroupWeight[x] * (self.expectedGroupUtilization[x] - self.groupUtilization[x]))
            
        '''


# In[ ]:

from numpy import *
class Eventlist:
    #currentTime = 0
    #collision = 0
    #maxTime = 0
    def __init__(self, endTime):
        self.event = []
        self.workList = []
        Global.maxTime = fix(endTime*1000000/52)
        self.nextTime = Global.rawInterval

        
    def findNextTimeEvents(self, points):
        self.event = []
        self.workList = []
        for pointCount in range(len(points)):
            if self.nextTime > points[pointCount].timeToNextTask:
                self.nextTime = points[pointCount].timeToNextTask
                self.event = [points[pointCount]]
            elif self.nextTime == points[pointCount].timeToNextTask:
                self.event.append(points[pointCount])
    
    def goToNextTime(self):
        Global.currentTime = self.nextTime
        self.nextTime = Global.rawInterval
        
    def changeState(self):
        for changeStateIndex in range(len(self.event)):
            self.event[changeStateIndex].changeState()
            #print "ID: %d , state: %d" % (self.event[x].ID, self.event[x].state)
            if self.event[changeStateIndex].state == Global.READY:
                self.workList.append(self.event[changeStateIndex])
    
    def timeToSendPacket(self, AP):
        if len(self.workList)>0:
            if self.workList[0].DEVICE == "access point":
                self.workList[0].sendPacket()
                del self.workList[0]            
        for timeToSendIndex in range(len(self.workList)):
           AP.receivePacket(self.workList[timeToSendIndex].sendPacket())
        
    def checkChannelState(self):
        if Global.channelState == Global.IDLE:
            if Global.channelUser > 0:
                Global.channelBusyStartTime = Global.currentTime
                Global.channelState = Global.carrierSense
        elif Global.channelState == Global.carrierSense:
            if Global.channelUser == 0:
                Global.channelUsingTime = Global.channelUsingTime + (Global.currentTime - Global.channelBusyStartTime)
                Global.channelState = Global.IDLE          
        #print "check channel: %d" % Global.channelUser
    
    def wakeUpGroup(self, group):
        for wakeIndex in range(1, len(group)):
            group[wakeIndex].wakeUp()
    
    def groupGoSleep(self, group):
        for sleepIndex in range(1, len(group)):
            group[sleepIndex].goSleep()
    
    def showList():
        for x in range(len(event)):
            print Eventlist.event[x]
    


# In[ ]:

from random import randint
class DataInterval:
    random.seed(6551)
    def __init__(self):
        self.arrivalTimeList = [Global.maxTime]
        self.nodeArrivalTime = [Global.maxTime]
        #self.dataAmount = [0]
        
    def checkDataArrival(self, points):
        for arrivalIndex in range(1, len(points)):
            if self.arrivalTimeList[arrivalIndex] < Global.currentTime:
                points[arrivalIndex].dataInterval()
                self.arrivalTimeList[arrivalIndex] = Global.currentTime + self.nodeArrivalTime[arrivalIndex]
                
    def getSamplingRate(self, points):
        for getSampleIndex in range(1, len(points)):
            self.nodeArrivalTime.append(int(((60 * 1000000) / Global.slotTime) / points[getSampleIndex].samplingRate))
            self.arrivalTimeList.append((Global.firstData.randint(0 ,self.nodeArrivalTime[getSampleIndex])))
            #self.dataAmount.append(1)


# In[ ]:

from __future__ import division
from random import choice, randint
class Global:
    '''constant'''
    SIFS = 3
    DIFS = 5
    ACKTime = 5
    slotTime = 52
    dataRate = 0.65    
    nav = SIFS + ACKTime
    contentionWindow = 4
    beacon_sec = 0.45            
    beaconPerSec = (1 / beacon_sec)
    beaconTime = fix((beacon_sec * 1000000) / slotTime)    
            
    RADIUS = 1000
    SQUERE_RADIUS = RADIUS ** 2
    packetLengthList = [64, 128, 256]
    samplingRateList = [2, 8, 16, 32] #per minute    
    '''constant'''
    
    '''no meaning, state ID'''
    carrierSense = 635
    sendPacket = 531 
    waitDIFS = 564 
    waitResponse = 315 #wait for ACK
    backoff = 615
    READY = 534
    HALT = 5843
    IDLE = 655
    '''no meaning, state ID'''
    
    
    

    '''SEED'''
    globalSeed = random
    globalSeed.seed(666)
    
    locationSeed = random
    beaconCountSeed = random
    firstData = random
    choiceSeed = random
    groupingSeed = random
    backoffSeed =random
    
    getSeed = []
    seedList = []
    for globalSeedIndex in range(1, 10000):
        getSeed.append(globalSeedIndex)
    for insertSeedIndex in range(100):
        insertSeed = []
        for appendSeed in range(6):
            insertSeed.append(getSeed.pop(getSeed.index(globalSeed.choice(getSeed))))
        seedList.append(insertSeed)
    '''SEED'''

    currentTime = 0
    maxTime = 0
    numOfNode = 256    
    group = 16
    rawTime = (beaconTime / group)
    holdingPeriod = (rawTime * 0.2)
    packetLength = 128
    
    
    '''變數'''
    nodeCounts = 0
    beaconInterval = currentTime + beaconTime
    rawInterval = currentTime + rawTime
    holdingPeriodInterval = rawInterval - holdingPeriod
    
    channelUser = 0
    channelBusyStartTime = 0
    channelUsingTime = 0
    channelState = IDLE    
        
    packetCount = 0
    collisionTimes = 0
    success = 0
    rawSlotCount = 0
    competitorCount = 0 #total
    extryCompetitorCount = 0 #total


# In[ ]:

from random import choice, randint
from numpy import *
class main:
    
    
    def __init__(self, nodeParameter):
        self.assignRateList = []
        self.points = []
        self.dataIntervalController = []
        self.eventController = []
        self.totalNodes = ((Global.numOfNode)*nodePara)
    
    def start(self, time, node_mode):
        for samplRate in range(len(Global.samplingRateList)):
            for singleType in range(int(self.totalNodes/len(Global.samplingRateList))):
                self.assignRateList.append(Global.samplingRateList[samplRate])

        self.eventController = Eventlist(time)
        self.dataIntervalController = DataInterval()
        self.points.append(AP(0,0,Global.group))


        while Global.nodeCounts < self.totalNodes:
            x, y = Global.locationSeed.randint(-Global.RADIUS,Global.RADIUS), Global.locationSeed.randint(-Global.RADIUS,Global.RADIUS)
            if (x*x + y*y) <= Global.SQUERE_RADIUS:
                self.points.append(node_mode(x, y, 128, self.assignRateList.pop(self.assignRateList.index(Global.choiceSeed.choice(self.assignRateList))), Global.nodeCounts))        

        #self.points[0].loadBalancedGrouping(self.points)
        self.points[0].randomGrouping(self.points)

        for calcRangeIndex in range(Global.group):
            for node1 in range(1, len(self.points[0].group[calcRangeIndex])):
                self.points[0].calcRange(self.points[0].group[calcRangeIndex][node1])
                for node2 in range(node1+1, len(self.points[0].group[calcRangeIndex])):
                    self.points[0].group[calcRangeIndex][node1].calcRange(self.points[0].group[calcRangeIndex][node2])

        self.dataIntervalController.getSamplingRate(self.points)

        self.lastRawTime = (Global.maxTime - Global.rawTime)
        while Global.currentTime < Global.maxTime:
            while Global.currentTime < min(self.lastRawTime, Global.beaconInterval):
                for groupTurn in range(Global.group):
                    self.dataIntervalController.checkDataArrival(self.points)
                    #wake up group x
                    self.eventController.wakeUpGroup(self.points[0].group[groupTurn])
                    while Global.currentTime < min(Global.maxTime, Global.rawInterval):
                        self.eventController.findNextTimeEvents(self.points[0].group[groupTurn])
                        self.eventController.goToNextTime()
                        self.eventController.changeState()
                        self.eventController.timeToSendPacket(self.points[0])
                        self.eventController.checkChannelState()
                    #group x go to sleep
                    self.eventController.groupGoSleep(self.points[0].group[groupTurn])
                    Global.rawInterval = Global.currentTime + Global.rawTime
                    Global.holdingPeriodInterval = Global.rawInterval - Global.holdingPeriod
                    Global.rawSlotCount += 1
            Global.beaconInterval = Global.currentTime + Global.beaconTime

        #print "end"
    
    def initialParameter(self, seedIndex):
        self.assignRateList = []
        self.points = []
        self.dataIntervalController = []
        self.eventController = []
        Global.nodeCounts = 0
        Global.currentTime = 0
        Global.beaconInterval = Global.beaconTime
        Global.rawInterval = Global.rawTime
        Global.holdingPeriodInterval = Global.rawInterval - Global.holdingPeriod
        Global.channelUser = 0
        Global.channelBusyStartTime = 0
        Global.channelUsingTime = 0
        Global.channelState = Global.IDLE 
        Global.groupingSeed.seed(Global.seedList[seedIndex][0])
        Global.choiceSeed.seed(Global.seedList[seedIndex][1])
        Global.locationSeed.seed(Global.seedList[seedIndex][2])
        Global.beaconCountSeed.seed(Global.seedList[seedIndex][3])
        Global.firstData.seed(Global.seedList[seedIndex][4])
        Global.backoffSeed.seed(Global.seedList[seedIndex][5])
        Global.packetCount = 0
        Global.collisionTimes = 0
        Global.success = 0
        Global.rawSlotCount = 0
        Global.competitorCount = 0 #total
        Global.extryCompetitorCount = 0 #total


# In[ ]:

from openpyxl import Workbook, load_workbook
from numpy import *
from __future__ import division
simTime = 80

wb = Workbook()
ws = wb.active
ws.append(["Nodes","Groups","Time","Send","Collision","Success","Collision probability","Throughput","Throughput(Kbps)","Utilization","Synchronous rate","Average compeptitor per raw slot", "Average extra competitor per raw slot", "method", "Random seed"])

#wb = load_workbook("simResult.xlsx")
#ws = wb.active
for nodePara in range(1, 9):
    IEEE80211ahSim = main(nodePara)
    #print nodePara
    for numOfTime in range(50):
        IEEE80211ahSim.initialParameter(numOfTime)
        IEEE80211ahSim.start(simTime, Node_common)
        collision_prob = round((Global.collisionTimes / Global.packetCount),5)
        throughput = (Global.success * Global.packetLength)
        throughput_kbps = round((((throughput * 8) / simTime) / 1000),2)
        utilization = round((Global.channelUsingTime / Global.maxTime),5)
        synchronousRate = 0
        for statistic in range(Global.group):
            for statistic2 in range(1, len(IEEE80211ahSim.points[0].group[statistic])):
                synchronousRate += (IEEE80211ahSim.points[0].group[statistic][statistic2].simultaneouslyCount / (IEEE80211ahSim.points[0].group[statistic][statistic2].totalSamplingTime -1))
        synchronousRate = round((synchronousRate / (len(IEEE80211ahSim.points)-1)), 5)
        avgCompetitorPerSlot = round((Global.competitorCount / Global.rawSlotCount), 5)
        avgExCompetitorPerSlot = round((Global.extryCompetitorCount / Global.rawSlotCount),5)
        ws.append([IEEE80211ahSim.totalNodes, Global.group, simTime, Global.packetCount, Global.collisionTimes, Global.success, collision_prob, throughput, throughput_kbps, utilization, synchronousRate,avgCompetitorPerSlot,avgExCompetitorPerSlot, "common", Global.seedList[numOfTime][0], Global.seedList[numOfTime][1], Global.seedList[numOfTime][2],Global.seedList[numOfTime][3], Global.seedList[numOfTime][4], Global.seedList[numOfTime][5]])
        print "in %d parameter, sim time: %d end" % (nodePara, numOfTime)
    print "nodes amount: %d, end" % (nodePara*Global.numOfNode)
print "process end: common method" 

for nodePara in range(1, 9):
    IEEE80211ahSim = main(nodePara)
    #print nodePara
    for numOfTime in range(50):
        IEEE80211ahSim.initialParameter(numOfTime)
        IEEE80211ahSim.start(simTime, Node_scheduled)
        collision_prob = round((Global.collisionTimes / Global.packetCount),5)
        throughput = (Global.success * Global.packetLength)
        throughput_kbps = round((((throughput * 8) / simTime) / 1000),2)
        utilization = round((Global.channelUsingTime / Global.maxTime),5)
        synchronousRate = 0
        for statistic in range(Global.group):
            for statistic2 in range(1, len(IEEE80211ahSim.points[0].group[statistic])):
                synchronousRate += (IEEE80211ahSim.points[0].group[statistic][statistic2].simultaneouslyCount / (IEEE80211ahSim.points[0].group[statistic][statistic2].totalSamplingTime -1))
        synchronousRate = round((synchronousRate / (len(IEEE80211ahSim.points)-1)), 5)
        avgCompetitorPerSlot = round((Global.competitorCount / Global.rawSlotCount), 5)
        avgExCompetitorPerSlot = round((Global.extryCompetitorCount / Global.rawSlotCount),5)
        ws.append([IEEE80211ahSim.totalNodes, Global.group, simTime, Global.packetCount, Global.collisionTimes, Global.success, collision_prob, throughput, throughput_kbps, utilization, synchronousRate,avgCompetitorPerSlot,avgExCompetitorPerSlot, "scheduled", Global.seedList[numOfTime][0], Global.seedList[numOfTime][1], Global.seedList[numOfTime][2],Global.seedList[numOfTime][3], Global.seedList[numOfTime][4], Global.seedList[numOfTime][5]])
        print "in %d parameter, sim time: %d end" % (nodePara, numOfTime)
    print "nodes amount: %d, end" % (nodePara*Global.numOfNode)
print "process end: scheduled method" 

wb.save("simResult_ver3.xlsx")


# In[ ]:

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")


# In[ ]:

#print Global.seedList
#print int(Global.numOfNode/Global.group)
#print len(points)
#print Global.numOfNode
#print Global.nodeCounts
#print len(IEEE80211ahSim.assignRateList)
#print IEEE80211ahSim.totalNodes
'''print len(IEEE80211ahSim.points)
for x in range(1,len(IEEE80211ahSim.points)):
    if IEEE80211ahSim.points[x].totalSamplingTime == 1:
        print IEEE80211ahSim.dataIntervalController.arrivalTimeList[x]'''
print (Global.success *128)
print Global.currentTime
print Global.beaconTime
print IEEE80211ahSim.lastRawTime
print Global.extryCompetitorCount


# In[ ]:



